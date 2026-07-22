"""Excel export for the member report history."""
from __future__ import annotations

import asyncio
import datetime as dt
import io
import logging
from collections import Counter
from typing import Any

import discord
from discord import app_commands
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from bot.cogs import voice_audit
from bot.cogs.report_system import ensure_schema

logger = logging.getLogger(__name__)
KST = dt.timezone(dt.timedelta(hours=9))
MAX_DISCORD_FILE_BYTES = 24 * 1024 * 1024

STATUS_LABELS = {
    "open": "접수됨",
    "claimed": "처리 중",
    "resolved": "처리 완료",
    "rejected": "신고 기각",
}
STATUS_FILLS = {
    "open": "FFF2CC",
    "claimed": "D9EAF7",
    "resolved": "D9EAD3",
    "rejected": "E7E6E6",
}


def is_staff(interaction: discord.Interaction) -> bool:
    member = interaction.user
    return isinstance(member, discord.Member) and (
        member.id == voice_audit.ADMIN_USER_ID
        or member.guild_permissions.manage_guild
        or member.guild_permissions.moderate_members
    )


def load_report_history(guild_id: int, days: int, status: str) -> list[dict[str, Any]]:
    ensure_schema()
    conditions = ["guild_id=%s"]
    params: list[Any] = [guild_id]

    if days > 0:
        conditions.append("created_at >= %s")
        params.append(voice_audit.now() - dt.timedelta(days=days))
    if status != "all":
        conditions.append("status=%s")
        params.append(status)

    query = f"""
        select id, guild_id, channel_id, message_id, reporter_id, target_id,
               target_text, reason, details, status, claimed_by, resolved_by,
               created_at, updated_at
        from public.discordbot_reports
        where {' and '.join(conditions)}
        order by created_at desc, id desc
    """
    with voice_audit.db() as conn, conn.cursor() as cur:
        cur.execute(query, tuple(params))
        return [dict(row) for row in cur.fetchall()]


def safe_excel_text(value: Any) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text[:32767]


def excel_datetime(value: Any) -> dt.datetime | None:
    if not isinstance(value, dt.datetime):
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(KST).replace(tzinfo=None)


def duration_text(created_at: Any, updated_at: Any, status: str) -> str:
    if status not in {"resolved", "rejected"}:
        return ""
    if not isinstance(created_at, dt.datetime) or not isinstance(updated_at, dt.datetime):
        return ""
    seconds = max(0, int((updated_at - created_at).total_seconds()))
    days, remainder = divmod(seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes = remainder // 60
    if days:
        return f"{days}일 {hours:02d}시간 {minutes:02d}분"
    if hours:
        return f"{hours}시간 {minutes:02d}분"
    return f"{minutes}분"


def member_name(name_map: dict[int, str], user_id: Any) -> str:
    if not user_id:
        return ""
    numeric_id = int(user_id)
    return safe_excel_text(name_map.get(numeric_id) or "서버에 없음")


def build_report_workbook(
    rows: list[dict[str, Any]],
    *,
    guild_id: int,
    guild_name: str,
    name_map: dict[int, str],
    days: int,
    status_filter: str,
) -> io.BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "요약"
    details = workbook.create_sheet("신고 내역")

    title_fill = PatternFill("solid", fgColor="8B1E2D")
    header_fill = PatternFill("solid", fgColor="B52B3A")
    white_bold = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="F4CCCC")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    summary.merge_cells("A1:F1")
    summary["A1"] = "멤버 신고 내역"
    summary["A1"].fill = title_fill
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30

    period_label = "전체 기간" if days == 0 else f"최근 {days}일"
    status_label = "전체 상태" if status_filter == "all" else STATUS_LABELS.get(status_filter, status_filter)
    generated_at = dt.datetime.now(KST).replace(tzinfo=None)
    metadata = [
        ("서버", safe_excel_text(guild_name)),
        ("서버 ID", str(guild_id)),
        ("출력 시각", generated_at),
        ("기간 필터", period_label),
        ("상태 필터", status_label),
        ("총 신고 건수", len(rows)),
    ]
    for index, (label, value) in enumerate(metadata, start=3):
        summary.cell(index, 1, label)
        summary.cell(index, 2, value)
        summary.cell(index, 1).font = Font(bold=True)
        summary.cell(index, 1).fill = section_fill
        summary.cell(index, 1).border = border
        summary.cell(index, 2).border = border
    summary["B5"].number_format = "yyyy-mm-dd hh:mm:ss"

    summary["A11"] = "상태"
    summary["B11"] = "건수"
    for cell in summary[11]:
        if cell.column <= 2:
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    counts = Counter(str(row.get("status") or "open") for row in rows)
    status_order = ["open", "claimed", "resolved", "rejected"]
    for row_index, status in enumerate(status_order, start=12):
        summary.cell(row_index, 1, STATUS_LABELS[status])
        summary.cell(row_index, 2, counts.get(status, 0))
        summary.cell(row_index, 1).border = border
        summary.cell(row_index, 2).border = border
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=STATUS_FILLS[status])

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "신고 처리 상태"
    chart.y_axis.title = "건수"
    chart.x_axis.title = "상태"
    data = Reference(summary, min_col=2, min_row=11, max_row=15)
    categories = Reference(summary, min_col=1, min_row=12, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 12
    summary.add_chart(chart, "D3")

    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 3
    summary.sheet_view.showGridLines = False

    headers = [
        "신고 번호",
        "처리 상태",
        "접수 시각",
        "최근 변경 시각",
        "처리 소요시간",
        "신고자",
        "신고자 ID",
        "신고 대상",
        "신고 대상 ID",
        "신고 사유",
        "상세 내용 및 증거",
        "담당자",
        "담당자 ID",
        "최종 처리자",
        "최종 처리자 ID",
        "신고 채널 ID",
        "신고 메시지 ID",
        "신고 메시지 링크",
    ]
    details.append(headers)

    for row in rows:
        reporter_id = int(row["reporter_id"]) if row.get("reporter_id") else None
        target_id = int(row["target_id"]) if row.get("target_id") else None
        claimed_by = int(row["claimed_by"]) if row.get("claimed_by") else None
        resolved_by = int(row["resolved_by"]) if row.get("resolved_by") else None
        channel_id = int(row["channel_id"]) if row.get("channel_id") else None
        message_id = int(row["message_id"]) if row.get("message_id") else None
        status = str(row.get("status") or "open")
        message_url = (
            f"https://discord.com/channels/{guild_id}/{channel_id}/{message_id}"
            if channel_id and message_id
            else ""
        )
        target_name = member_name(name_map, target_id) if target_id else safe_excel_text(row.get("target_text"))

        details.append(
            [
                int(row["id"]),
                STATUS_LABELS.get(status, status),
                excel_datetime(row.get("created_at")),
                excel_datetime(row.get("updated_at")),
                duration_text(row.get("created_at"), row.get("updated_at"), status),
                member_name(name_map, reporter_id),
                reporter_id,
                target_name,
                target_id,
                safe_excel_text(row.get("reason")),
                safe_excel_text(row.get("details")),
                member_name(name_map, claimed_by),
                claimed_by,
                member_name(name_map, resolved_by),
                resolved_by,
                channel_id,
                message_id,
                message_url,
            ]
        )

    for cell in details[1]:
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    details.sheet_view.showGridLines = False
    details.row_dimensions[1].height = 27

    widths = [12, 14, 20, 20, 18, 22, 22, 24, 22, 36, 65, 22, 22, 22, 22, 22, 22, 60]
    for column_index, width in enumerate(widths, start=1):
        details.column_dimensions[get_column_letter(column_index)].width = width

    if rows:
        table = Table(displayName="ReportHistoryTable", ref=f"A1:R{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        details.add_table(table)

    for row_index, row in enumerate(rows, start=2):
        status = str(row.get("status") or "open")
        details.cell(row_index, 2).fill = PatternFill("solid", fgColor=STATUS_FILLS.get(status, "FFFFFF"))
        details.cell(row_index, 3).number_format = "yyyy-mm-dd hh:mm:ss"
        details.cell(row_index, 4).number_format = "yyyy-mm-dd hh:mm:ss"
        details.cell(row_index, 18).style = "Hyperlink"
        if details.cell(row_index, 18).value:
            details.cell(row_index, 18).hyperlink = details.cell(row_index, 18).value
        for column_index in range(1, 19):
            cell = details.cell(row_index, column_index)
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column_index in {8, 10, 11, 18},
            )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


class ReportExportCog(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot

    @app_commands.command(name="신고엑셀", description="[관리자] 신고 내역을 엑셀 파일로 출력합니다.")
    @app_commands.describe(period="출력할 신고 기간", status="출력할 처리 상태")
    @app_commands.choices(
        period=[
            app_commands.Choice(name="전체 기간", value=0),
            app_commands.Choice(name="최근 7일", value=7),
            app_commands.Choice(name="최근 30일", value=30),
            app_commands.Choice(name="최근 90일", value=90),
            app_commands.Choice(name="최근 1년", value=365),
        ],
        status=[
            app_commands.Choice(name="전체 상태", value="all"),
            app_commands.Choice(name="접수됨", value="open"),
            app_commands.Choice(name="처리 중", value="claimed"),
            app_commands.Choice(name="처리 완료", value="resolved"),
            app_commands.Choice(name="신고 기각", value="rejected"),
        ],
    )
    async def export_reports(
        self,
        interaction: discord.Interaction,
        period: app_commands.Choice[int] | None = None,
        status: app_commands.Choice[str] | None = None,
    ) -> None:
        if interaction.guild is None:
            await interaction.response.send_message("서버 안에서만 사용할 수 있습니다.", ephemeral=True)
            return
        if not is_staff(interaction):
            await interaction.response.send_message("운영진만 신고 내역을 출력할 수 있습니다.", ephemeral=True)
            return

        await interaction.response.defer(ephemeral=True, thinking=True)
        days = period.value if period else 0
        status_filter = status.value if status else "all"

        try:
            rows = await asyncio.to_thread(
                load_report_history,
                interaction.guild.id,
                days,
                status_filter,
            )
            if not rows:
                await interaction.followup.send(
                    "선택한 조건에 해당하는 신고 내역이 없습니다.",
                    ephemeral=True,
                )
                return

            user_ids: set[int] = set()
            for row in rows:
                for key in ("reporter_id", "target_id", "claimed_by", "resolved_by"):
                    value = row.get(key)
                    if value:
                        user_ids.add(int(value))
            name_map = {
                user_id: member.display_name
                for user_id in user_ids
                if (member := interaction.guild.get_member(user_id)) is not None
            }

            output = await asyncio.to_thread(
                build_report_workbook,
                rows,
                guild_id=interaction.guild.id,
                guild_name=interaction.guild.name,
                name_map=name_map,
                days=days,
                status_filter=status_filter,
            )
            file_size = output.getbuffer().nbytes
            if file_size > MAX_DISCORD_FILE_BYTES:
                await interaction.followup.send(
                    "엑셀 파일이 Discord 첨부 한도를 넘었습니다. 기간이나 상태를 좁혀 다시 출력해 주세요.",
                    ephemeral=True,
                )
                return

            timestamp = dt.datetime.now(KST).strftime("%Y%m%d_%H%M")
            filename = f"report_history_{interaction.guild.id}_{timestamp}.xlsx"
            period_label = "전체 기간" if days == 0 else f"최근 {days}일"
            status_label = "전체 상태" if status_filter == "all" else STATUS_LABELS[status_filter]
            await interaction.followup.send(
                content=(
                    f"✅ 신고 내역 **{len(rows)}건**을 출력했습니다.\n"
                    f"기간: **{period_label}** · 상태: **{status_label}**"
                ),
                file=discord.File(output, filename=filename),
                ephemeral=True,
            )
        except Exception as exc:
            logger.exception("신고 내역 엑셀 출력 실패", exc_info=exc)
            await interaction.followup.send(
                f"❌ 신고 내역 엑셀 출력 실패: `{type(exc).__name__}`",
                ephemeral=True,
            )


async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(ReportExportCog(bot))
