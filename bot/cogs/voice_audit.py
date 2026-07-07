"""Supabase PostgreSQL 기반 전체 멤버 음성 활동 및 미접속 경고 관리."""
from __future__ import annotations

import asyncio
import datetime as dt
import io
import logging
import math
import os
from zoneinfo import ZoneInfo

import discord
import psycopg
from discord import app_commands
from discord.ext import commands, tasks
from openpyxl import Workbook
from psycopg.rows import dict_row

logger = logging.getLogger(__name__)
ADMIN_USER_ID = 324558739921305602
WARNING_DAYS = 7
KST = ZoneInfo("Asia/Seoul")
FEATURE_BASELINE = dt.datetime(2026, 7, 1, 0, 0, tzinfo=KST)
FIRST_REPORT_AT = dt.datetime(2026, 7, 8, 10, 0, tzinfo=KST)
DATABASE_URL = os.getenv("SUPABASE_DB_URL")


def now() -> dt.datetime:
    return dt.datetime.now(KST)


def db() -> psycopg.Connection:
    if not DATABASE_URL:
        raise RuntimeError("SUPABASE_DB_URL 환경변수가 없습니다")
    return psycopg.connect(
        DATABASE_URL,
        row_factory=dict_row,
        connect_timeout=10,
        prepare_threshold=None,
    )


def parse_time(value):
    if value is None:
        return None
    parsed = value if isinstance(value, dt.datetime) else dt.datetime.fromisoformat(value)
    return parsed.astimezone(KST) if parsed.tzinfo else parsed.replace(tzinfo=KST)


def member_joined_at(member: discord.Member) -> dt.datetime:
    joined = member.joined_at or now()
    return parse_time(joined)


def tracking_baseline(member: discord.Member) -> dt.datetime:
    return max(FEATURE_BASELINE, member_joined_at(member))


def duration_text(seconds: int) -> str:
    days, rem = divmod(max(0, seconds), 86400)
    hours, rem = divmod(rem, 3600)
    minutes = rem // 60
    parts = []
    if days:
        parts.append(f"{days}일")
    if hours:
        parts.append(f"{hours}시간")
    parts.append(f"{minutes}분")
    return " ".join(parts)


def verify_database() -> str:
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute("select current_database() as name")
            return str(cur.fetchone()["name"])


def upsert_members(guild_id: int, members: list[discord.Member]) -> None:
    current = now()
    rows = [
        (guild_id, member.id, member.display_name, tracking_baseline(member), member_joined_at(member), current)
        for member in members
        if not member.bot
    ]
    if not rows:
        return
    with db() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                insert into public.discordbot_member_activity
                (guild_id,user_id,display_name,baseline_at,joined_at,updated_at)
                values(%s,%s,%s,%s,%s,%s)
                on conflict(guild_id,user_id)
                do update set
                  display_name=excluded.display_name,
                  baseline_at=excluded.baseline_at,
                  joined_at=excluded.joined_at,
                  updated_at=excluded.updated_at
                """,
                rows,
            )


def load_activity(guild_id: int) -> dict[int, dict]:
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "select * from public.discordbot_member_activity where guild_id=%s",
                (guild_id,),
            )
            return {int(row["user_id"]): row for row in cur.fetchall()}


def build_activity_rows(guild_id: int, members: list[discord.Member]) -> list[dict]:
    upsert_members(guild_id, members)
    saved = load_activity(guild_id)
    current = now()
    rows = []
    for member in members:
        if member.bot:
            continue
        row = saved.get(member.id)
        baseline = parse_time(row["baseline_at"]) if row else tracking_baseline(member)
        last_voice = parse_time(row["last_voice_at"]) if row else None
        reference = last_voice or baseline or current
        rows.append({
            "member": member,
            "last_voice": last_voice,
            "inactive_seconds": max(0, int((current - reference).total_seconds())),
            "warnings": int(row["warning_count"]) if row else 0,
            "joined_at": member_joined_at(member),
        })
    return sorted(rows, key=lambda item: item["inactive_seconds"], reverse=True)


def authorized(interaction: discord.Interaction) -> bool:
    return interaction.user.id == ADMIN_USER_ID


async def deny(interaction: discord.Interaction) -> None:
    text = "이 기능은 지정된 관리자만 사용할 수 있습니다."
    if interaction.response.is_done():
        await interaction.followup.send(text, ephemeral=True)
    else:
        await interaction.response.send_message(text, ephemeral=True)


async def db_error(interaction: discord.Interaction, exc: Exception) -> None:
    logger.exception("음성 관리 DB 작업 실패", exc_info=exc)
    await interaction.followup.send(
        f"❌ 음성 관리 작업 실패\n오류: `{type(exc).__name__}`\n"
        "Server Members Intent와 SUPABASE_DB_URL을 확인해주세요.",
        ephemeral=True,
    )


class VoiceAuditView(discord.ui.View):
    def __init__(self, cog: "VoiceAuditCog"):
        super().__init__(timeout=600)
        self.cog = cog

    async def interaction_check(self, interaction: discord.Interaction) -> bool:
        if authorized(interaction):
            return True
        await deny(interaction)
        return False

    @discord.ui.button(label="전체 멤버 현황", emoji="👥", style=discord.ButtonStyle.primary)
    async def members(self, interaction: discord.Interaction, _: discord.ui.Button):
        await interaction.response.defer(ephemeral=True, thinking=True)
        try:
            members = await self.cog.fetch_all_members(interaction.guild)
            embed = await asyncio.to_thread(self.cog.member_embed, interaction.guild, members)
            await interaction.followup.send(embed=embed, ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @discord.ui.button(label="경고 현황", emoji="⚠️", style=discord.ButtonStyle.danger)
    async def warning_status(self, interaction: discord.Interaction, _: discord.ui.Button):
        await interaction.response.defer(ephemeral=True, thinking=True)
        try:
            members = await self.cog.fetch_all_members(interaction.guild)
            embed = await asyncio.to_thread(self.cog.warning_embed, interaction.guild, members)
            await interaction.followup.send(embed=embed, ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @discord.ui.button(label="최근 입퇴장", emoji="🎙️", style=discord.ButtonStyle.secondary)
    async def recent(self, interaction: discord.Interaction, _: discord.ui.Button):
        await interaction.response.defer(ephemeral=True)
        try:
            embed = await asyncio.to_thread(self.cog.recent_embed, interaction.guild)
            await interaction.followup.send(embed=embed, ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @discord.ui.button(label="즉시 경고 점검", emoji="🔍", style=discord.ButtonStyle.success)
    async def warnings(self, interaction: discord.Interaction, _: discord.ui.Button):
        await interaction.response.defer(ephemeral=True, thinking=True)
        try:
            members = await self.cog.fetch_all_members(interaction.guild)
            awarded = await asyncio.to_thread(self.cog.check_warnings, interaction.guild, members, 1)
            await interaction.followup.send(f"새 경고 **{len(awarded)}명 / {sum(item['count'] for item in awarded)}회**를 부여했습니다.", ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @discord.ui.button(label="Excel 출력", emoji="📊", style=discord.ButtonStyle.success)
    async def excel(self, interaction: discord.Interaction, _: discord.ui.Button):
        await interaction.response.defer(ephemeral=True, thinking=True)
        try:
            members = await self.cog.fetch_all_members(interaction.guild)
            file = await asyncio.to_thread(self.cog.make_excel, interaction.guild, members)
            await interaction.followup.send(file=file, ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)


class VoiceAuditCog(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.weekly_report_loop.start()

    def cog_unload(self):
        self.weekly_report_loop.cancel()

    async def fetch_all_members(self, guild: discord.Guild) -> list[discord.Member]:
        if not self.bot.intents.members:
            raise RuntimeError("Server Members Intent가 비활성화되어 있습니다")
        members = [member async for member in guild.fetch_members(limit=None)]
        return [member for member in members if not member.bot]

    @commands.Cog.listener()
    async def on_ready(self):
        for guild in self.bot.guilds:
            try:
                members = await self.fetch_all_members(guild)
                await asyncio.to_thread(upsert_members, guild.id, members)
                await asyncio.to_thread(self.reconcile_voice_state, guild)
                logger.info(
                    "전체 멤버 음성 추적 초기화: guild=%s fetched=%s member_count=%s",
                    guild.id,
                    len(members),
                    guild.member_count,
                )
            except Exception:
                logger.exception("전체 멤버 초기화 실패: %s", guild.id)

    @commands.Cog.listener()
    async def on_member_join(self, member: discord.Member):
        if member.bot:
            return
        try:
            await asyncio.to_thread(upsert_members, member.guild.id, [member])
        except Exception:
            logger.exception("신규 멤버 등록 실패: %s", member.id)

    def reconcile_voice_state(self, guild: discord.Guild) -> None:
        current = now()
        current_members = {
            member.id: (member, channel)
            for channel in guild.voice_channels
            for member in channel.members
            if not member.bot
        }
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "select * from public.discordbot_active_voice_sessions where guild_id=%s",
                    (guild.id,),
                )
                active_rows = {int(row["user_id"]): row for row in cur.fetchall()}
                for user_id, row in active_rows.items():
                    if user_id in current_members:
                        continue
                    joined_at = parse_time(row["joined_at"])
                    cur.execute(
                        """
                        insert into public.discordbot_voice_sessions
                        (guild_id,user_id,display_name,channel_id,channel_name,joined_at,left_at,duration_seconds)
                        values(%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            guild.id,user_id,row["display_name"],row["channel_id"],row["channel_name"],
                            joined_at,current,max(0,int((current-joined_at).total_seconds())),
                        ),
                    )
                    cur.execute(
                        "delete from public.discordbot_active_voice_sessions where guild_id=%s and user_id=%s",
                        (guild.id,user_id),
                    )
                for user_id, (member, channel) in current_members.items():
                    existing = active_rows.get(user_id)
                    joined_at = parse_time(existing["joined_at"]) if existing else current
                    cur.execute(
                        """
                        insert into public.discordbot_active_voice_sessions
                        (guild_id,user_id,display_name,channel_id,channel_name,joined_at,updated_at)
                        values(%s,%s,%s,%s,%s,%s,%s)
                        on conflict(guild_id,user_id)
                        do update set display_name=excluded.display_name,channel_id=excluded.channel_id,
                        channel_name=excluded.channel_name,updated_at=excluded.updated_at
                        """,
                        (guild.id,user_id,member.display_name,channel.id,channel.name,joined_at,current),
                    )
                    cur.execute(
                        """
                        update public.discordbot_member_activity
                        set last_voice_at=%s,updated_at=%s
                        where guild_id=%s and user_id=%s
                        """,
                        (current,current,guild.id,user_id),
                    )

    @commands.Cog.listener()
    async def on_voice_state_update(self, member, before, after):
        if member.bot or before.channel == after.channel:
            return
        try:
            await asyncio.to_thread(self.record_voice_change, member, before, after)
        except Exception:
            logger.exception("음성 상태 기록 실패: %s", member.id)

    def record_voice_change(self, member, before, after):
        upsert_members(member.guild.id, [member])
        current = now()
        with db() as conn:
            with conn.cursor() as cur:
                if before.channel is not None:
                    cur.execute(
                        """
                        select * from public.discordbot_active_voice_sessions
                        where guild_id=%s and user_id=%s for update
                        """,
                        (member.guild.id,member.id),
                    )
                    active = cur.fetchone()
                    if active:
                        joined_at = parse_time(active["joined_at"])
                        cur.execute(
                            """
                            insert into public.discordbot_voice_sessions
                            (guild_id,user_id,display_name,channel_id,channel_name,joined_at,left_at,duration_seconds)
                            values(%s,%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                member.guild.id,member.id,member.display_name,
                                active["channel_id"],active["channel_name"],joined_at,current,
                                max(0,int((current-joined_at).total_seconds())),
                            ),
                        )
                        cur.execute(
                            "delete from public.discordbot_active_voice_sessions where guild_id=%s and user_id=%s",
                            (member.guild.id,member.id),
                        )
                if after.channel is not None:
                    cur.execute(
                        """
                        insert into public.discordbot_active_voice_sessions
                        (guild_id,user_id,display_name,channel_id,channel_name,joined_at,updated_at)
                        values(%s,%s,%s,%s,%s,%s,%s)
                        on conflict(guild_id,user_id)
                        do update set display_name=excluded.display_name,channel_id=excluded.channel_id,
                        channel_name=excluded.channel_name,joined_at=excluded.joined_at,updated_at=excluded.updated_at
                        """,
                        (member.guild.id,member.id,member.display_name,after.channel.id,after.channel.name,current,current),
                    )
                cur.execute(
                    """
                    update public.discordbot_member_activity
                    set display_name=%s,last_voice_at=%s,updated_at=%s
                    where guild_id=%s and user_id=%s
                    """,
                    (member.display_name,current,current,member.guild.id,member.id),
                )

    def check_warnings(self, guild: discord.Guild, members: list[discord.Member], max_per_member: int = 1) -> list[dict]:
        upsert_members(guild.id, members)
        current = now()
        member_map = {member.id: member for member in members}
        awarded: list[dict] = []
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute("select pg_try_advisory_xact_lock(%s) as locked", (guild.id,))
                if not cur.fetchone()["locked"]:
                    return []
                cur.execute(
                    "select * from public.discordbot_member_activity where guild_id=%s for update",
                    (guild.id,),
                )
                for row in cur.fetchall():
                    user_id = int(row["user_id"])
                    member = member_map.get(user_id)
                    if member is None:
                        continue
                    joined_at = member_joined_at(member)
                    if (current - joined_at).total_seconds() < WARNING_DAYS * 86400:
                        continue
                    baseline = parse_time(row["baseline_at"]) or tracking_baseline(member)
                    last_voice = parse_time(row["last_voice_at"])
                    last_warning = parse_time(row["last_warning_at"])
                    reference = max(value for value in (baseline,last_voice,last_warning) if value)
                    count = math.floor((current-reference).total_seconds()/(WARNING_DAYS*86400))
                    count = min(count, max_per_member)
                    if count <= 0:
                        continue
                    warning_time = reference + dt.timedelta(days=WARNING_DAYS*count)
                    new_total = int(row["warning_count"]) + count
                    cur.execute(
                        """
                        update public.discordbot_member_activity
                        set warning_count=warning_count+%s,last_warning_at=%s,updated_at=%s
                        where guild_id=%s and user_id=%s
                        """,
                        (count,warning_time,current,guild.id,user_id),
                    )
                    for index in range(count):
                        awarded_at = reference + dt.timedelta(days=WARNING_DAYS*(index+1))
                        cur.execute(
                            """
                            insert into public.discordbot_warning_history
                            (guild_id,user_id,display_name,awarded_at,inactivity_days,reason)
                            values(%s,%s,%s,%s,%s,%s)
                            """,
                            (guild.id,user_id,member.display_name,awarded_at,WARNING_DAYS,"음성채널 7일 미접속"),
                        )
                    awarded.append({"member": member, "count": count, "total": new_total})
        return awarded

    def get_report_channel_id(self, guild_id: int) -> int | None:
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute("select channel_id from public.discordbot_voice_report_config where guild_id=%s", (guild_id,))
                row = cur.fetchone()
                return int(row["channel_id"]) if row and row["channel_id"] else None

    def set_report_channel_id(self, guild_id: int, channel_id: int) -> None:
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    insert into public.discordbot_voice_report_config(guild_id,channel_id,updated_at)
                    values(%s,%s,%s)
                    on conflict(guild_id)
                    do update set channel_id=excluded.channel_id,updated_at=excluded.updated_at
                    """,
                    (guild_id,channel_id,now()),
                )

    def reserve_report(self, guild_id: int, report_key: str, week: int) -> bool:
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    insert into public.discordbot_voice_audit_report_log
                    (guild_id,report_key,report_week,reported_at,warning_member_count)
                    values(%s,%s,%s,%s,0)
                    on conflict(guild_id,report_key) do nothing
                    """,
                    (guild_id,report_key,week,now()),
                )
                return cur.rowcount == 1

    def finish_report(self, guild_id: int, report_key: str, channel_id: int, warning_member_count: int) -> None:
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    update public.discordbot_voice_audit_report_log
                    set recipient_channel_id=%s,warning_member_count=%s,reported_at=%s
                    where guild_id=%s and report_key=%s
                    """,
                    (channel_id,warning_member_count,now(),guild_id,report_key),
                )

    def build_report_embeds(self, guild: discord.Guild, week: int, awarded: list[dict]) -> list[discord.Embed]:
        report_date = now().strftime("%Y-%m-%d %H:%M")
        title = f"⚠️ 음성 미접속 경고자 명단 · {week}주차"
        if not awarded:
            embed = discord.Embed(title=title, color=0x57F287)
            embed.description = "이번 주 새로 경고가 부여된 멤버가 없습니다."
            embed.set_footer(text=f"기준: 7일 이상 음성채널 미접속 · {report_date} KST")
            return [embed]
        lines = [
            f"{idx}. {item['member'].mention} · **{item['member'].display_name}** · 누적 경고 **{item['total']}회**"
            for idx, item in enumerate(awarded, start=1)
        ]
        chunks: list[list[str]] = []
        chunk: list[str] = []
        length = 0
        for line in lines:
            if chunk and length + len(line) + 1 > 3800:
                chunks.append(chunk)
                chunk = []
                length = 0
            chunk.append(line)
            length += len(line) + 1
        if chunk:
            chunks.append(chunk)
        embeds = []
        for page, chunk_lines in enumerate(chunks, start=1):
            embed = discord.Embed(title=title, color=0xED4245)
            embed.description = "\n".join(chunk_lines)
            embed.add_field(name="이번 주 신규 경고", value=f"{len(awarded)}명", inline=True)
            embed.add_field(name="기준", value="7일 이상 음성채널 미접속", inline=True)
            embed.set_footer(text=f"{page}/{len(chunks)} · {report_date} KST")
            embeds.append(embed)
        return embeds

    async def send_weekly_report(self, guild: discord.Guild, report_key: str, week: int) -> None:
        channel_id = await asyncio.to_thread(self.get_report_channel_id, guild.id)
        if channel_id is None:
            logger.warning("음성 경고 보고 채널 미설정: guild=%s", guild.id)
            return
        channel = guild.get_channel(channel_id) or await self.bot.fetch_channel(channel_id)
        if not isinstance(channel, (discord.TextChannel, discord.Thread)):
            logger.warning("음성 경고 보고 채널이 텍스트 채널이 아님: guild=%s channel=%s", guild.id, channel_id)
            return
        members = await self.fetch_all_members(guild)
        awarded = await asyncio.to_thread(self.check_warnings, guild, members, 1)
        embeds = self.build_report_embeds(guild, week, awarded)
        for embed in embeds:
            await channel.send(embed=embed)
        await asyncio.to_thread(self.finish_report, guild.id, report_key, channel_id, len(awarded))

    def panel_embed(self, guild: discord.Guild, members: list[discord.Member], database_name: str) -> discord.Embed:
        rows = build_activity_rows(guild.id, members)
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute("select count(*) as count from public.discordbot_voice_sessions where guild_id=%s", (guild.id,))
                sessions = int(cur.fetchone()["count"])
                cur.execute("select count(*) as count from public.discordbot_active_voice_sessions where guild_id=%s", (guild.id,))
                active = int(cur.fetchone()["count"])
                cur.execute("select channel_id from public.discordbot_voice_report_config where guild_id=%s", (guild.id,))
                report_row = cur.fetchone()
        report_channel = f"<#{report_row['channel_id']}>" if report_row and report_row["channel_id"] else "미설정"
        embed = discord.Embed(title="🎙️ 음성 활동 관리 패널", color=0x57F287)
        embed.description = "✅ **Supabase DB 연결 정상**\n✅ **전체 멤버 목록 불러오기 정상**"
        embed.add_field(name="프로젝트", value="kokiriko", inline=True)
        embed.add_field(name="전체 멤버", value=f"{len(members)}명", inline=True)
        embed.add_field(name="현재 음성 접속", value=f"{active}명", inline=True)
        embed.add_field(name="경고 공지 채널", value=report_channel, inline=True)
        embed.add_field(name="7일 이상 미접속", value=f"{sum(r['inactive_seconds'] >= 604800 for r in rows)}명", inline=True)
        embed.add_field(name="경고 보유", value=f"{sum(r['warnings'] > 0 for r in rows)}명", inline=True)
        embed.add_field(name="완료된 세션", value=f"{sessions:,}건", inline=True)
        embed.set_footer(text=f"DB: {database_name} · 확인 {now().strftime('%Y-%m-%d %H:%M:%S')} KST")
        return embed

    def member_embed(self, guild: discord.Guild, members: list[discord.Member]) -> discord.Embed:
        rows = build_activity_rows(guild.id, members)
        lines = [
            f"{row['member'].mention} · 미접속 **{duration_text(row['inactive_seconds'])}** · 경고 **{row['warnings']}회**"
            for row in rows[:35]
        ]
        embed = discord.Embed(title="👥 전체 멤버 미접속 및 경고", color=0x3498DB)
        embed.description = "\n".join(lines)[:4000] or "표시할 멤버가 없습니다."
        embed.set_footer(text=f"전체 {len(rows)}명 중 35명 표시 · 전체 목록은 Excel에서 확인")
        return embed

    def warning_embed(self, guild: discord.Guild, members: list[discord.Member]) -> discord.Embed:
        rows = [row for row in build_activity_rows(guild.id, members) if row["warnings"] > 0]
        embed = discord.Embed(title="⚠️ 누적 경고 현황", color=0xED4245)
        embed.description = "\n".join(
            f"{row['member'].mention} · 경고 **{row['warnings']}회** · 미접속 {duration_text(row['inactive_seconds'])}"
            for row in rows[:35]
        )[:4000] or "경고 보유 멤버가 없습니다."
        return embed

    def recent_embed(self, guild: discord.Guild) -> discord.Embed:
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "select * from public.discordbot_voice_sessions where guild_id=%s order by id desc limit 20",
                    (guild.id,),
                )
                rows = cur.fetchall()
        embed = discord.Embed(title="🎙️ 최근 음성채널 입퇴장 기록", color=0x57F287)
        embed.description = "\n".join(
            f"<@{row['user_id']}> · **{row['channel_name']}** · {duration_text(row['duration_seconds'])}"
            for row in rows
        ) or "기록된 세션이 없습니다."
        return embed

    def make_excel(self, guild: discord.Guild, members: list[discord.Member]) -> discord.File:
        activity = build_activity_rows(guild.id, members)
        book = Workbook()
        status_sheet = book.active
        status_sheet.title = "전체 멤버 미접속 및 경고"
        status_sheet.append(["유저 ID","닉네임","가입일","최근 음성 접속","미접속 시간","미접속 일수","누적 경고"])
        for row in activity:
            status_sheet.append([
                row["member"].id,
                row["member"].display_name,
                str(row["joined_at"]),
                str(row["last_voice"] or "기능 적용 후 접속 기록 없음"),
                duration_text(row["inactive_seconds"]),
                round(row["inactive_seconds"]/86400,2),
                row["warnings"],
            ])
        sessions_sheet = book.create_sheet("음성 입퇴장 기록")
        sessions_sheet.append(["유저 ID","닉네임","채널명","입장","퇴장","사용 시간(분)"])
        warning_sheet = book.create_sheet("경고 이력")
        warning_sheet.append(["유저 ID","닉네임","부여 시각","기준 일수","사유"])
        with db() as conn:
            with conn.cursor() as cur:
                cur.execute("select * from public.discordbot_voice_sessions where guild_id=%s order by joined_at desc", (guild.id,))
                for row in cur.fetchall():
                    sessions_sheet.append([
                        row["user_id"],row["display_name"],row["channel_name"],
                        str(row["joined_at"]),str(row["left_at"]),round(row["duration_seconds"]/60,1),
                    ])
                cur.execute("select * from public.discordbot_warning_history where guild_id=%s order by awarded_at desc", (guild.id,))
                for row in cur.fetchall():
                    warning_sheet.append([
                        row["user_id"],row["display_name"],str(row["awarded_at"]),row["inactivity_days"],row["reason"],
                    ])
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        return discord.File(output, filename=f"voice_audit_{now().strftime('%Y%m%d_%H%M')}.xlsx")

    @app_commands.command(name="음성관리패널", description="[전용 관리자] 전체 멤버 음성 미접속 현황을 관리합니다.")
    async def panel(self, interaction: discord.Interaction):
        if not authorized(interaction):
            await deny(interaction)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        try:
            members = await self.fetch_all_members(interaction.guild)
            database_name = await asyncio.to_thread(verify_database)
            embed = await asyncio.to_thread(self.panel_embed, interaction.guild, members, database_name)
            await interaction.followup.send(embed=embed, view=VoiceAuditView(self), ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @app_commands.command(name="음성기록엑셀", description="[전용 관리자] 전체 멤버 음성 기록을 Excel로 출력합니다.")
    async def export_excel(self, interaction: discord.Interaction):
        if not authorized(interaction):
            await deny(interaction)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        try:
            members = await self.fetch_all_members(interaction.guild)
            file = await asyncio.to_thread(self.make_excel, interaction.guild, members)
            await interaction.followup.send(file=file, ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @app_commands.command(name="음성공지채널설정", description="[전용 관리자] 주간 음성 미접속 경고 명단을 보낼 채널을 설정합니다.")
    @app_commands.describe(channel="경고자 명단을 보낼 공지 채널")
    async def set_report_channel(self, interaction: discord.Interaction, channel: discord.TextChannel):
        if not authorized(interaction):
            await deny(interaction)
            return
        await interaction.response.defer(ephemeral=True)
        try:
            await asyncio.to_thread(self.set_report_channel_id, interaction.guild.id, channel.id)
            await interaction.followup.send(f"✅ 음성 경고자 명단 공지 채널을 {channel.mention}(으)로 설정했습니다.", ephemeral=True)
        except Exception as exc:
            await db_error(interaction, exc)

    @tasks.loop(minutes=5)
    async def weekly_report_loop(self):
        current = now()
        if current < FIRST_REPORT_AT:
            return
        if current.weekday() != 2 or current.hour != 10:
            return
        week = ((current.date() - FEATURE_BASELINE.date()).days // 7)
        if week < 1:
            return
        report_key = f"voice-warning-week-{week}-{current.date().isoformat()}"
        for guild in self.bot.guilds:
            try:
                reserved = await asyncio.to_thread(self.reserve_report, guild.id, report_key, week)
                if not reserved:
                    continue
                await self.send_weekly_report(guild, report_key, week)
            except Exception:
                logger.exception("주간 음성 경고 보고 실패: %s", guild.id)

    @weekly_report_loop.before_loop
    async def before_weekly_report_loop(self):
        await self.bot.wait_until_ready()


async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(VoiceAuditCog(bot))
